import psycopg2
from psycopg2.extras import RealDictCursor
import csv
import re
import os
import sys
import subprocess
import time
import socket
from dotenv import load_dotenv
from contextlib import contextmanager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================
# SELEÇÃO DE CLIENTE
# ============================================
def listar_clientes():
    """Lista todos os arquivos .env de clientes com seus nomes amigáveis."""
    from dotenv import dotenv_values
    
    clientes = []
    for arquivo in sorted(os.listdir('.')):
        if arquivo.startswith('.env.') and not arquivo.endswith('-git'):
            try:
                config = dotenv_values(arquivo)
                nome_cliente = config.get('NOME_CLIENTE', arquivo.replace('.env.', '').upper())
                clientes.append((nome_cliente, arquivo))
            except Exception as e:
                nome_fallback = arquivo.replace('.env.', '').upper()
                clientes.append((nome_fallback, arquivo))
                print(f"⚠️  Aviso: Não foi possível ler NOME_CLIENTE de {arquivo}")
    
    return clientes

def exibir_menu_clientes():
    """Exibe menu e retorna o arquivo .env selecionado."""
    clientes = listar_clientes()
    
    if not clientes:
        print("\n" + "="*60)
        print("❌ ERRO: Nenhum arquivo de configuração encontrado!")
        print("="*60)
        print("\nCrie arquivos no formato:")
        print("  .env.staging")
        print("\nUse o arquivo .env-git como template.")
        print("="*60)
        sys.exit(1)
    
    print("\n" + "="*60)
    print(" 🔧  SELEÇÃO DE CLIENTE - AJUSTE DE INCONSISTÊNCIAS")
    print("="*60)
    
    for idx, (nome, _) in enumerate(clientes, 1):
        print(f"  {idx}. {nome}")
    
    print("  0. Sair")
    print("="*60)
    
    while True:
        try:
            escolha = input("\n➤ Selecione o cliente (número): ").strip()
            
            if escolha == '0':
                print("\n⚠️  Operação cancelada pelo usuário.\n")
                sys.exit(0)
            
            idx = int(escolha) - 1
            
            if 0 <= idx < len(clientes):
                nome, arquivo = clientes[idx]
                print(f"\n✅ Cliente selecionado: {nome}")
                print("="*60)
                return arquivo, nome
            else:
                print("❌ Opção inválida! Tente novamente.")
        
        except ValueError:
            print("❌ Digite um número válido!")
        except KeyboardInterrupt:
            print("\n\n⚠️  Operação cancelada.\n")
            sys.exit(0)

# Seleciona o cliente e carrega as variáveis de ambiente
env_file, NOME_CLIENTE_SELECIONADO = exibir_menu_clientes()
load_dotenv(env_file)

# --- FUNÇÕES DE CONFIGURAÇÃO ---
def carregar_configuracoes():
    """Carrega configurações do arquivo .env atual."""
    # Host único para todos os bancos (otimização)
    db_host = os.getenv('DB_HOST', 'localhost')
    
    DB_GESTAO = {
        'host': db_host,
        'database': os.getenv('DB_GESTAO_NAME'),
        'user': os.getenv('DB_GESTAO_USER'),
        'password': os.getenv('DB_GESTAO_PASS')
    }

    DB_CONTRATO = {
        'host': db_host,
        'database': os.getenv('DB_CONTRATO_NAME'),
        'user': os.getenv('DB_CONTRATO_USER'),
        'password': os.getenv('DB_CONTRATO_PASS')
    }

    DB_PESSOA = {
        'host': db_host,
        'database': os.getenv('DB_PESSOA_NAME'),
        'user': os.getenv('DB_PESSOA_USER'),
        'password': os.getenv('DB_PESSOA_PASS')
    }

    SENHA_ACCOUNTS = os.getenv('DB_ACCOUNTS_PASS')
    URL_ACCOUNTS = os.getenv('URL_ACCOUNTS')
    DB_ACCOUNTS_NAME_USER = os.getenv('DB_ACCOUNTS_NAME_USER')

    SSH_CONFIG = {
        'ssh_host': os.getenv('SSH_HOST'),
        'ssh_user': os.getenv('SSH_USER'),
        'ssh_port': int(os.getenv('SSH_PORT', '22')),
        'ssh_password': os.getenv('SSH_PASSWORD'),
        'ssh_pkey': os.getenv('SSH_PKEY_PATH'),
        'remote_bind_address': (os.getenv('SSH_REMOTE_DB_HOST', 'localhost'), int(os.getenv('SSH_REMOTE_DB_PORT', '5432'))),
        'local_bind_port': int(os.getenv('SSH_LOCAL_PORT', '5435'))
    }
    
    # Validação: túnel SSH é obrigatório
    if not SSH_CONFIG['ssh_host'] or not SSH_CONFIG['ssh_user']:
        raise ValueError("SSH_HOST e SSH_USER são obrigatórios no arquivo .env")
    
    LIMITE_REGISTROS = int(os.getenv('LIMITE_REGISTROS', '0'))
    
    return DB_GESTAO, DB_CONTRATO, DB_PESSOA, SSH_CONFIG, SENHA_ACCOUNTS, URL_ACCOUNTS, DB_ACCOUNTS_NAME_USER, LIMITE_REGISTROS

# --- FUNÇÕES AUXILIARES ---
def limpar_cpf(cpf):
    """Remove caracteres não numéricos."""
    if not cpf: return None
    return re.sub(r'\D', '', str(cpf))

def formatar_cpf(cpf):
    """Aplica máscara de CPF."""
    c = limpar_cpf(cpf)
    if not c or len(c) != 11: return c
    return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}"

def comparar_campos(dict1, dict2, campos):
    """
    Compara campos entre dois dicionários.
    Retorna dict com campos divergentes: {'campo': (valor1, valor2)}
    """
    divergencias = {}
    for campo in campos:
        val1 = dict1.get(campo)
        val2 = dict2.get(campo)
        
        # Normaliza valores para comparação
        if campo == 'cpf_cnpj' or campo == 'cpf':
            val1 = limpar_cpf(val1)
            val2 = limpar_cpf(val2)
        elif isinstance(val1, str):
            val1 = val1.strip() if val1 else None
        if isinstance(val2, str):
            val2 = val2.strip() if val2 else None
            
        if val1 != val2:
            divergencias[campo] = (val1, val2)
    
    return divergencias

def verificar_porta_disponivel(port):
    """Verifica se uma porta está disponível para uso."""
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.bind(('127.0.0.1', port))
        sock.close()
        return True
    except OSError:
        return False

def aguardar_porta_aberta(port, timeout=10):
    """Aguarda até que a porta esteja aberta e aceitando conexões."""
    inicio = time.time()
    while time.time() - inicio < timeout:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        try:
            sock.connect(('127.0.0.1', port))
            sock.close()
            return True
        except (socket.error, ConnectionRefusedError):
            time.sleep(0.5)
    return False

@contextmanager
def gerenciar_tunnel_ssh(SSH_CONFIG):
    """Context manager para gerenciar ciclo de vida do túnel SSH."""
    processo_ssh = None
    
    try:
        print(f"[SSH] Conectando ao servidor {SSH_CONFIG['ssh_host']}:{SSH_CONFIG['ssh_port']}...")
        
        if not verificar_porta_disponivel(SSH_CONFIG['local_bind_port']):
            print(f"[SSH] Aviso: Porta {SSH_CONFIG['local_bind_port']} já está em uso.")
            print(f"[SSH] Assumindo que o túnel já está ativo...")
            yield None
            return
        
        remote_host, remote_port = SSH_CONFIG['remote_bind_address']
        
        ssh_cmd = [
            'ssh',
            '-L', f"{SSH_CONFIG['local_bind_port']}:{remote_host}:{remote_port}",
            '-p', str(SSH_CONFIG['ssh_port']),
            '-l', SSH_CONFIG['ssh_user'],
            SSH_CONFIG['ssh_host'],
            '-N',
            '-o', 'StrictHostKeyChecking=no',
            '-o', 'ServerAliveInterval=60',
            '-o', 'ServerAliveCountMax=3'
        ]
        
        if SSH_CONFIG['ssh_pkey']:
            ssh_cmd.insert(1, '-i')
            ssh_cmd.insert(2, SSH_CONFIG['ssh_pkey'])
        
        print(f"[SSH] Estabelecendo túnel: localhost:{SSH_CONFIG['local_bind_port']} -> {remote_host}:{remote_port}")
        
        if os.name == 'nt':
            processo_ssh = subprocess.Popen(
                ssh_cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                stdin=subprocess.PIPE,
                creationflags=subprocess.CREATE_NEW_PROCESS_GROUP
            )
        else:
            processo_ssh = subprocess.Popen(
                ssh_cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                stdin=subprocess.PIPE,
                preexec_fn=os.setsid
            )
        
        if SSH_CONFIG['ssh_password'] and not SSH_CONFIG['ssh_pkey']:
            print("[SSH] Nota: Para autenticação por senha, considere usar chave SSH.")
        
        print(f"[SSH] Aguardando túnel ficar ativo...", end=" ")
        if aguardar_porta_aberta(SSH_CONFIG['local_bind_port'], timeout=15):
            print("✓")
            print(f"[SSH] Túnel SSH estabelecido com sucesso!")
        else:
            raise Exception("Timeout ao aguardar túnel SSH ficar ativo")
        
        yield processo_ssh
        
    except FileNotFoundError:
        print(f"[SSH] ERRO: Comando 'ssh' não encontrado no sistema.")
        print(f"[SSH] Certifique-se de que o OpenSSH está instalado.")
        sys.exit(1)
    except Exception as e:
        print(f"[SSH] Erro ao estabelecer túnel: {e}")
        if processo_ssh:
            try:
                processo_ssh.terminate()
            except:
                pass
        sys.exit(1)
    finally:
        if processo_ssh:
            print("[SSH] Encerrando túnel SSH...", end=" ")
            try:
                processo_ssh.terminate()
                processo_ssh.wait(timeout=5)
                print("✓")
            except:
                try:
                    processo_ssh.kill()
                    print("✓ (forçado)")
                except:
                    print("⚠️  (processo pode continuar em background)")
            print("[SSH] Túnel SSH encerrado.")

def ajustar_hosts_para_tunnel(db_config, SSH_CONFIG):
    """Ajusta host e porta dos bancos para usar túnel SSH."""
    config = db_config.copy()
    config['host'] = '127.0.0.1'
    config['port'] = SSH_CONFIG['local_bind_port']
    return config

def ler_relatorio_emails_duplicados(cliente_nome):
    """Lê o relatório de emails duplicados gerado pelo script de análise."""
    nome_arquivo = f'relatorio_{cliente_nome.lower().replace(" ", "_")}.xlsx'
    caminho = os.path.join(os.getcwd(), '..', 'analise-inconsistencia', nome_arquivo)
    
    # Tenta encontrar o arquivo em diferentes localizações
    locais_possiveis = [
        caminho,
        os.path.join(os.getcwd(), nome_arquivo),
        os.path.join(os.getcwd(), '..', nome_arquivo),
    ]
    
    arquivo_encontrado = None
    for local in locais_possiveis:
        if os.path.exists(local):
            arquivo_encontrado = local
            break
    
    if not arquivo_encontrado:
        print(f"\n❌ ERRO: Relatório não encontrado!")
        print(f"\nArquivo procurado: {nome_arquivo}")
        print(f"\nLocais verificados:")
        for local in locais_possiveis:
            print(f"  - {local}")
        print(f"\n💡 Execute primeiro o script de análise para gerar o relatório.")
        sys.exit(1)
    
    print(f"[Relatório] Lendo: {arquivo_encontrado}")
    
    try:
        wb = load_workbook(arquivo_encontrado)
        ws = wb['1-Emails Duplicados']
        
        registros = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Se tem UUID
                registro = dict(zip(headers, row))
                registros.append(registro)
        
        wb.close()
        print(f"[Relatório] {len(registros)} registros carregados")
        return registros
        
    except Exception as e:
        print(f"❌ Erro ao ler relatório: {e}")
        sys.exit(1)

def salvar_excel_consolidado(relatorios_dict, nome_arquivo='ajuste_executado.xlsx'):
    """Salva múltiplos relatórios em um único arquivo Excel com abas separadas."""
    caminho = os.path.join(os.getcwd(), nome_arquivo)
    
    try:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        for nome_aba, (dados, cabecalho) in relatorios_dict.items():
            ws = wb.create_sheet(title=nome_aba)
            
            if not dados:
                ws.append(cabecalho)
                ws.append(['Nenhum registro encontrado'])
                continue
            
            ws.append(cabecalho)
            
            # Estiliza cabeçalho
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, _ in enumerate(cabecalho, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Adiciona dados
            for item in dados:
                linha = [item.get(col, '') for col in cabecalho]
                ws.append(linha)
            
            # Ajusta largura das colunas
            for col_num, col_name in enumerate(cabecalho, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(str(col_name))
                for row in ws.iter_rows(min_row=2, max_row=min(100, len(dados)+1), min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            ws.freeze_panes = 'A2'
        
        wb.save(caminho)
        
        total_registros = sum(len(dados) for dados, _ in relatorios_dict.values())
        print(f"\n📊 Relatório de execução salvo: {caminho}")
        print(f"   └─ {len(relatorios_dict)} abas criadas | {total_registros} registros totais")
        
    except Exception as e:
        print(f"⚠️  Erro ao salvar arquivo Excel: {e}")

def main():
    # Carrega configurações
    try:
        DB_GESTAO, DB_CONTRATO, DB_PESSOA, SSH_CONFIG, SENHA_ACCOUNTS, URL_ACCOUNTS, DB_ACCOUNTS_NAME_USER, LIMITE_REGISTROS = carregar_configuracoes()
    except ValueError as e:
        print(f"❌ ERRO: {e}")
        sys.exit(1)
    
    cliente_nome = os.getenv('NOME_CLIENTE', 'CLIENTE')
    
    print(f"\n--- INICIANDO AJUSTE DE INCONSISTÊNCIAS [{cliente_nome}] ---")
    
    # Gerencia túnel SSH automaticamente
    with gerenciar_tunnel_ssh(SSH_CONFIG):
        # Ajusta configurações dos bancos para usar túnel
        db_gestao_ajustado = ajustar_hosts_para_tunnel(DB_GESTAO, SSH_CONFIG)
        db_contrato_ajustado = ajustar_hosts_para_tunnel(DB_CONTRATO, SSH_CONFIG)
        
        # Lê relatório de emails duplicados
        print("\n" + "="*60)
        print("ETAPA 1: CARREGAMENTO DE DADOS")
        print("="*60)
        
        registros = ler_relatorio_emails_duplicados(cliente_nome)
        
        if not registros:
            print("❌ Nenhum registro de email duplicado encontrado no relatório.")
            return
        
        # Aplica limite de registros se configurado
        MODO_DEBUG = False
        if LIMITE_REGISTROS > 0:
            registros = registros[:LIMITE_REGISTROS]
            if LIMITE_REGISTROS == 1:
                MODO_DEBUG = True
                print(f"🔍 MODO DEBUG ATIVADO: Processamento interativo detalhado")
            else:
                print(f"⚠️  LIMITE ATIVO: Processando apenas {len(registros)} registros (LIMITE_REGISTROS={LIMITE_REGISTROS})")
        else:
            print(f"📊 Processando todos os {len(registros)} registros")
        
        # Listas para relatório final
        lista_updates_gestao = []
        lista_updates_contrato = []
        lista_desvinculacoes = []
        lista_erros = []
        lista_ignorados = []
        
        contador_processados = 0
        contador_atualizados_gestao = 0
        contador_atualizados_contrato = 0
        contador_desvinculados = 0
        
        print("\n" + "="*60)
        print("ETAPA 2: ANÁLISE E PREPARAÇÃO DE UPDATES")
        print("="*60)
        
        # Conexões com os bancos
        try:
            conn_gestao = psycopg2.connect(**db_gestao_ajustado)
            conn_contrato = psycopg2.connect(**db_contrato_ajustado)
            
            print("[Conexões] Bancos conectados com sucesso!")
            
            # Processa cada registro
            for idx, registro in enumerate(registros, 1):
                uuid = registro['uuid_comum']
                
                if MODO_DEBUG:
                    print("\n" + "="*70)
                    print(f"🔍 ANÁLISE DETALHADA - REGISTRO {idx}/{len(registros)}")
                    print("="*70)
                    print(f"UUID: {uuid}")
                else:
                    print(f"\n[{idx}/{len(registros)}] Processando UUID: {uuid}")
                
                try:
                    # 1. Buscar dados em accounts (via dblink)
                    cur_gestao = conn_gestao.cursor(cursor_factory=RealDictCursor)
                    sql_accounts = f"""
                        SELECT id, cpf_cnpj, name, email, phone
                        FROM dblink(
                            'host={URL_ACCOUNTS} dbname={DB_ACCOUNTS_NAME_USER} user={DB_ACCOUNTS_NAME_USER} password={SENHA_ACCOUNTS}',
                            'SELECT id, cpf_cnpj, name, email, phone FROM users WHERE id = ''{uuid}'''
                        ) AS accounts(id uuid, cpf_cnpj varchar, name varchar, email varchar, phone varchar)
                    """
                    cur_gestao.execute(sql_accounts)
                    dados_accounts = cur_gestao.fetchone()
                    
                    if not dados_accounts:
                        print(f"  ⚠️  UUID não encontrado em accounts - IGNORANDO")
                        lista_ignorados.append({
                            'uuid': uuid,
                            'motivo': 'UUID não encontrado em accounts'
                        })
                        continue
                    
                    cpf_accounts = limpar_cpf(dados_accounts['cpf_cnpj'])
                    if not cpf_accounts:
                        print(f"  ⚠️  CPF vazio em accounts - IGNORANDO")
                        lista_ignorados.append({
                            'uuid': uuid,
                            'motivo': 'CPF vazio em accounts'
                        })
                        continue
                    
                    if MODO_DEBUG:
                        print(f"\n📋 DADOS EM ACCOUNTS (Fonte da Verdade):")
                        print(f"   CPF......: {formatar_cpf(cpf_accounts)}")
                        print(f"   Nome.....: {dados_accounts['name']}")
                        print(f"   Email....: {dados_accounts['email']}")
                        print(f"   Telefone.: {dados_accounts['phone'] or 'N/A'}")
                    else:
                        print(f"  ✓ Accounts: CPF={formatar_cpf(cpf_accounts)}, Nome={dados_accounts['name']}")
                    
                    # 2. Verificar existência em segurado (por CPF)
                    cur_contrato = conn_contrato.cursor(cursor_factory=RealDictCursor)
                    sql_segurado = """
                        SELECT id, cpf_cnpj, usuario_id, nome
                        FROM segurado
                        WHERE REGEXP_REPLACE(cpf_cnpj, '\D', '', 'g') = %s
                        LIMIT 1
                    """
                    cur_contrato.execute(sql_segurado, (cpf_accounts,))
                    dados_segurado = cur_contrato.fetchone()
                    
                    if not dados_segurado:
                        print(f"  ⚠️  CPF não encontrado em segurado - IGNORANDO")
                        lista_ignorados.append({
                            'uuid': uuid,
                            'cpf_accounts': formatar_cpf(cpf_accounts),
                            'motivo': 'CPF não encontrado em segurado'
                        })
                        continue
                    
                    if MODO_DEBUG:
                        print(f"\n✅ VALIDAÇÃO: CPF existe em SEGURADO")
                        print(f"   Segurado ID: {dados_segurado['id']}")
                        print(f"   Nome.......: {dados_segurado['nome']}")
                    else:
                        print(f"  ✓ Segurado encontrado: ID={dados_segurado['id']}")
                    
                    # 3. Comparar e preparar update para gestao.tb_usuario
                    sql_gestao_busca = """
                        SELECT id, cpf_cnpj, name, email, phone
                        FROM tb_usuario
                        WHERE sso_id = %s
                    """
                    cur_gestao.execute(sql_gestao_busca, (uuid,))
                    dados_gestao = cur_gestao.fetchone()
                    
                    if dados_gestao:
                        campos_comparar = ['cpf_cnpj', 'name', 'email', 'phone']
                        divergencias_gestao = comparar_campos(
                            {'cpf_cnpj': cpf_accounts, 'name': dados_accounts['name'], 
                             'email': dados_accounts['email'], 'phone': dados_accounts['phone']},
                            dict(dados_gestao),
                            campos_comparar
                        )
                        
                        if divergencias_gestao:
                            if MODO_DEBUG:
                                print(f"\n⚠️  DIVERGÊNCIAS EM GESTÃO.TB_USUARIO:")
                                for campo, (val_correto, val_atual) in divergencias_gestao.items():
                                    campo_label = {
                                        'cpf_cnpj': 'CPF',
                                        'name': 'Nome',
                                        'email': 'Email',
                                        'phone': 'Telefone'
                                    }.get(campo, campo)
                                    
                                    if campo == 'cpf_cnpj':
                                        val_correto = formatar_cpf(val_correto)
                                        val_atual = formatar_cpf(val_atual)
                                    
                                    print(f"   {campo_label}:")
                                    print(f"      Atual.....: {val_atual or 'N/A'}")
                                    print(f"      Correto...: {val_correto or 'N/A'}")
                            else:
                                print(f"  → Gestão: {len(divergencias_gestao)} campo(s) divergente(s)")
                            
                            lista_updates_gestao.append({
                                'uuid': uuid,
                                'id_gestao': dados_gestao['id'],
                                'cpf_antes': formatar_cpf(dados_gestao['cpf_cnpj']),
                                'cpf_depois': formatar_cpf(cpf_accounts),
                                'nome_antes': dados_gestao['name'],
                                'nome_depois': dados_accounts['name'],
                                'email_antes': dados_gestao['email'],
                                'email_depois': dados_accounts['email'],
                                'phone_antes': dados_gestao['phone'],
                                'phone_depois': dados_accounts['phone'],
                                'divergencias': str(list(divergencias_gestao.keys()))
                            })
                            contador_atualizados_gestao += 1
                        else:
                            if MODO_DEBUG:
                                print(f"\n✅ GESTÃO.TB_USUARIO: Dados consistentes")
                            else:
                                print(f"  ✓ Gestão: Dados consistentes")
                    
                    # 4. Comparar e preparar update para contrato.usuario
                    sql_contrato_busca = """
                        SELECT id, cpf_cnpj, nome, email
                        FROM usuario
                        WHERE sso_id = %s
                    """
                    cur_contrato.execute(sql_contrato_busca, (uuid,))
                    dados_contrato_usuario = cur_contrato.fetchone()
                    
                    if dados_contrato_usuario:
                        campos_comparar = ['cpf_cnpj', 'nome', 'email']
                        divergencias_contrato = comparar_campos(
                            {'cpf_cnpj': cpf_accounts, 'nome': dados_accounts['name'], 
                             'email': dados_accounts['email']},
                            {'cpf_cnpj': dados_contrato_usuario['cpf_cnpj'],
                             'nome': dados_contrato_usuario['nome'],
                             'email': dados_contrato_usuario['email']},
                            campos_comparar
                        )
                        
                        if divergencias_contrato:
                            if MODO_DEBUG:
                                print(f"\n⚠️  DIVERGÊNCIAS EM CONTRATO.USUARIO:")
                                for campo, (val_correto, val_atual) in divergencias_contrato.items():
                                    campo_label = {
                                        'cpf_cnpj': 'CPF',
                                        'nome': 'Nome',
                                        'email': 'Email'
                                    }.get(campo, campo)
                                    
                                    if campo == 'cpf_cnpj':
                                        val_correto = formatar_cpf(val_correto)
                                        val_atual = formatar_cpf(val_atual)
                                    
                                    print(f"   {campo_label}:")
                                    print(f"      Atual.....: {val_atual or 'N/A'}")
                                    print(f"      Correto...: {val_correto or 'N/A'}")
                            else:
                                print(f"  → Contrato.usuario: {len(divergencias_contrato)} campo(s) divergente(s)")
                            
                            lista_updates_contrato.append({
                                'uuid': uuid,
                                'id_usuario': dados_contrato_usuario['id'],
                                'cpf_antes': formatar_cpf(dados_contrato_usuario['cpf_cnpj']),
                                'cpf_depois': formatar_cpf(cpf_accounts),
                                'nome_antes': dados_contrato_usuario['nome'],
                                'nome_depois': dados_accounts['name'],
                                'email_antes': dados_contrato_usuario['email'],
                                'email_depois': dados_accounts['email'],
                                'divergencias': str(list(divergencias_contrato.keys()))
                            })
                            contador_atualizados_contrato += 1
                        else:
                            if MODO_DEBUG:
                                print(f"\n✅ CONTRATO.USUARIO: Dados consistentes")
                            else:
                                print(f"  ✓ Contrato.usuario: Dados consistentes")
                        
                        # 5. Verificar segurados com CPF divergente vinculados a este usuario_id
                        usuario_id = dados_contrato_usuario['id']
                        sql_segurados_divergentes = """
                            SELECT id, cpf_cnpj, nome
                            FROM segurado
                            WHERE usuario_id = %s
                            AND REGEXP_REPLACE(cpf_cnpj, '\D', '', 'g') != %s
                        """
                        cur_contrato.execute(sql_segurados_divergentes, (usuario_id, cpf_accounts))
                        segurados_divergentes = cur_contrato.fetchall()
                        
                        if segurados_divergentes:
                            if MODO_DEBUG:
                                print(f"\n⚠️  SEGURADOS COM CPF DIVERGENTE (serão desvinculados):")
                                for seg in segurados_divergentes:
                                    print(f"   Segurado ID: {seg['id']}")
                                    print(f"   CPF Errado.: {seg['cpf_cnpj']}")
                                    print(f"   CPF Correto: {formatar_cpf(cpf_accounts)}")
                                    print(f"   Nome.......: {seg['nome']}")
                                    print(f"   Ação.......: SET usuario_id = NULL")
                                    print()
                            else:
                                print(f"  → {len(segurados_divergentes)} segurado(s) com CPF divergente para desvincular")
                            
                            for seg in segurados_divergentes:
                                lista_desvinculacoes.append({
                                    'uuid': uuid,
                                    'segurado_id': seg['id'],
                                    'cpf_segurado': seg['cpf_cnpj'],
                                    'cpf_correto': formatar_cpf(cpf_accounts),
                                    'nome_segurado': seg['nome'],
                                    'usuario_id': usuario_id
                                })
                                contador_desvinculados += 1
                    
                    contador_processados += 1
                    
                    # Em modo debug, pausa após cada registro
                    if MODO_DEBUG:
                        print("\n" + "="*70)
                        print("📊 RESUMO DAS AÇÕES PARA ESTE REGISTRO:")
                        if lista_updates_gestao and lista_updates_gestao[-1]['uuid'] == uuid:
                            print("   ✓ UPDATE em gestao.tb_usuario")
                        if lista_updates_contrato and lista_updates_contrato[-1]['uuid'] == uuid:
                            print("   ✓ UPDATE em contrato.usuario")
                        if any(d['uuid'] == uuid for d in lista_desvinculacoes):
                            count = sum(1 for d in lista_desvinculacoes if d['uuid'] == uuid)
                            print(f"   ✓ Desvincular {count} segurado(s)")
                        if not lista_updates_gestao and not lista_updates_contrato and not any(d['uuid'] == uuid for d in lista_desvinculacoes):
                            print("   ✅ Nenhuma alteração necessária - Dados consistentes!")
                        print("="*70)
                    
                except Exception as e:
                    print(f"  ❌ Erro ao processar: {e}")
                    lista_erros.append({
                        'uuid': uuid,
                        'erro': str(e)
                    })
            
            # Resumo antes da execução
            print("\n" + "="*60)
            print("RESUMO DAS ALTERAÇÕES A SEREM EXECUTADAS")
            print("="*60)
            print(f"Registros processados: {contador_processados}")
            print(f"  - Updates em gestao.tb_usuario: {contador_atualizados_gestao}")
            print(f"  - Updates em contrato.usuario: {contador_atualizados_contrato}")
            print(f"  - Desvinculações em segurado: {contador_desvinculados}")
            print(f"  - Registros ignorados: {len(lista_ignorados)}")
            print(f"  - Erros: {len(lista_erros)}")
            print("="*60)
            
            # Confirmação do usuário
            if contador_atualizados_gestao == 0 and contador_atualizados_contrato == 0 and contador_desvinculados == 0:
                print("\n✅ Nenhuma alteração necessária! Todos os dados estão consistentes.")
                conn_gestao.close()
                conn_contrato.close()
                return
            
            if MODO_DEBUG:
                print("\n" + "="*70)
                print("🔍 MODO DEBUG - CONFIRMAÇÃO DETALHADA")
                print("="*70)
                print("Você revisou todas as divergências acima.")
                print("As alterações estão CORRETAS e serão aplicadas no banco.")
                print("="*70)
            else:
                print("\n⚠️  ATENÇÃO: As alterações serão executadas DIRETAMENTE no banco de dados!")
            
            resposta = input("\nConfirmar execução dos UPDATEs? (S/N): ").strip().upper()
            
            if resposta not in ['S', 'SIM', 'Y', 'YES']:
                print("\n⚠️  Operação cancelada pelo usuário.")
                conn_gestao.close()
                conn_contrato.close()
                return
            
            # Execução dos UPDATEs
            print("\n" + "="*60)
            print("ETAPA 3: EXECUÇÃO DOS UPDATES")
            print("="*60)
            
            # Updates em gestao.tb_usuario
            if lista_updates_gestao:
                print(f"\n[Gestão] Executando {len(lista_updates_gestao)} update(s)...")
                cur_gestao = conn_gestao.cursor()
                for item in lista_updates_gestao:
                    try:
                        sql_update = """
                            UPDATE tb_usuario
                            SET cpf_cnpj = %s, name = %s, email = %s, phone = %s, updated_at = NOW()
                            WHERE sso_id = %s
                        """
                        cur_gestao.execute(sql_update, (
                            item['cpf_depois'].replace('.', '').replace('-', ''),
                            item['nome_depois'],
                            item['email_depois'],
                            item['phone_depois'],
                            item['uuid']
                        ))
                        item['status'] = 'SUCESSO'
                    except Exception as e:
                        item['status'] = f'ERRO: {e}'
                        print(f"  ❌ Erro ao atualizar UUID {item['uuid']}: {e}")
                
                conn_gestao.commit()
                print(f"  ✓ Updates em gestao.tb_usuario concluídos")
            
            # Updates em contrato.usuario
            if lista_updates_contrato:
                print(f"\n[Contrato] Executando {len(lista_updates_contrato)} update(s)...")
                cur_contrato = conn_contrato.cursor()
                for item in lista_updates_contrato:
                    try:
                        sql_update = """
                            UPDATE usuario
                            SET cpf_cnpj = %s, nome = %s, email = %s, updated_at = NOW()
                            WHERE sso_id = %s
                        """
                        cur_contrato.execute(sql_update, (
                            item['cpf_depois'].replace('.', '').replace('-', ''),
                            item['nome_depois'],
                            item['email_depois'],
                            item['uuid']
                        ))
                        item['status'] = 'SUCESSO'
                    except Exception as e:
                        item['status'] = f'ERRO: {e}'
                        print(f"  ❌ Erro ao atualizar UUID {item['uuid']}: {e}")
                
                conn_contrato.commit()
                print(f"  ✓ Updates em contrato.usuario concluídos")
            
            # Desvinculações em segurado
            if lista_desvinculacoes:
                print(f"\n[Segurado] Executando {len(lista_desvinculacoes)} desvinculação(ões)...")
                cur_contrato = conn_contrato.cursor()
                for item in lista_desvinculacoes:
                    try:
                        sql_update = """
                            UPDATE segurado
                            SET usuario_id = NULL, updated_at = NOW()
                            WHERE id = %s
                        """
                        cur_contrato.execute(sql_update, (item['segurado_id'],))
                        item['status'] = 'SUCESSO'
                    except Exception as e:
                        item['status'] = f'ERRO: {e}'
                        print(f"  ❌ Erro ao desvincular segurado {item['segurado_id']}: {e}")
                
                conn_contrato.commit()
                print(f"  ✓ Desvinculações em segurado concluídas")
            
            # Em modo debug, valida os dados após update
            if MODO_DEBUG and contador_processados > 0:
                print("\n" + "="*70)
                print("🔍 VALIDAÇÃO PÓS-EXECUÇÃO")
                print("="*70)
                
                uuid_validar = registros[0]['uuid_comum']
                
                # Re-busca dados atualizados
                cur_gestao = conn_gestao.cursor(cursor_factory=RealDictCursor)
                cur_gestao.execute("SELECT cpf_cnpj, name, email, phone FROM tb_usuario WHERE sso_id = %s", (uuid_validar,))
                dados_gestao_apos = cur_gestao.fetchone()
                
                cur_contrato = conn_contrato.cursor(cursor_factory=RealDictCursor)
                cur_contrato.execute("SELECT cpf_cnpj, nome, email FROM usuario WHERE sso_id = %s", (uuid_validar,))
                dados_contrato_apos = cur_contrato.fetchone()
                
                print(f"\n📋 DADOS ATUALIZADOS EM GESTÃO:")
                print(f"   CPF......: {formatar_cpf(dados_gestao_apos['cpf_cnpj'])}")
                print(f"   Nome.....: {dados_gestao_apos['name']}")
                print(f"   Email....: {dados_gestao_apos['email']}")
                print(f"   Telefone.: {dados_gestao_apos['phone'] or 'N/A'}")
                
                print(f"\n📋 DADOS ATUALIZADOS EM CONTRATO:")
                print(f"   CPF......: {formatar_cpf(dados_contrato_apos['cpf_cnpj'])}")
                print(f"   Nome.....: {dados_contrato_apos['nome']}")
                print(f"   Email....: {dados_contrato_apos['email']}")
                
                print("\n✅ Validação concluída!")
                print("="*70)
            
            conn_gestao.close()
            conn_contrato.close()
            
            print("\n✅ Todas as alterações foram executadas com sucesso!")
            
            # Gerar relatório de execução
            print("\n" + "="*60)
            print("ETAPA 4: GERANDO RELATÓRIO DE EXECUÇÃO")
            print("="*60)
            
            # Headers para cada aba
            headers_resumo = ['Métrica', 'Valor']
            headers_gestao = ['uuid', 'id_gestao', 'cpf_antes', 'cpf_depois', 'nome_antes', 'nome_depois', 
                             'email_antes', 'email_depois', 'phone_antes', 'phone_depois', 'divergencias', 'status']
            headers_contrato = ['uuid', 'id_usuario', 'cpf_antes', 'cpf_depois', 'nome_antes', 'nome_depois',
                               'email_antes', 'email_depois', 'divergencias', 'status']
            headers_desvinc = ['uuid', 'segurado_id', 'cpf_segurado', 'cpf_correto', 'nome_segurado', 'usuario_id', 'status']
            headers_ignorados = ['uuid', 'cpf_accounts', 'motivo']
            headers_erros = ['uuid', 'erro']
            
            # Dados do resumo
            dados_resumo = [
                {'Métrica': 'Total de registros processados', 'Valor': contador_processados},
                {'Métrica': 'Updates em gestao.tb_usuario', 'Valor': contador_atualizados_gestao},
                {'Métrica': 'Updates em contrato.usuario', 'Valor': contador_atualizados_contrato},
                {'Métrica': 'Desvinculações em segurado', 'Valor': contador_desvinculados},
                {'Métrica': 'Registros ignorados', 'Valor': len(lista_ignorados)},
                {'Métrica': 'Erros encontrados', 'Valor': len(lista_erros)},
                {'Métrica': 'Cliente', 'Valor': cliente_nome},
                {'Métrica': 'Data/Hora', 'Valor': time.strftime('%Y-%m-%d %H:%M:%S')}
            ]
            
            relatorios = {
                '0-Resumo': (dados_resumo, headers_resumo),
                '1-Updates Gestão': (lista_updates_gestao, headers_gestao),
                '2-Updates Contrato': (lista_updates_contrato, headers_contrato),
                '3-Desvinculações': (lista_desvinculacoes, headers_desvinc),
                '4-Ignorados': (lista_ignorados, headers_ignorados),
                '5-Erros': (lista_erros, headers_erros)
            }
            
            nome_arquivo_relatorio = f'ajuste_executado_{cliente_nome.lower().replace(" ", "_")}.xlsx'
            salvar_excel_consolidado(relatorios, nome_arquivo_relatorio)
            
            print("\n" + "="*60)
            print("✅ AJUSTE DE INCONSISTÊNCIAS CONCLUÍDO COM SUCESSO!")
            print("="*60)
            
        except Exception as e:
            print(f"\n❌ Erro crítico: {e}")
            print("⚠️  Verifique as conexões e tente novamente.")
            return

if __name__ == "__main__":
    main()

